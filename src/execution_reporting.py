# -*- coding: utf-8 -*-
# Script details
# Created On 16-may-2019
# Develop By: Abheet

import os
from src.logger import *
from openpyxl import load_workbook
import openpyxl
from src.read_settings import *
from src.dashboard import format_output_file


class ReportGenerator():
    def __init__(self,record):
        self.rb = openpyxl.Workbook()
        self.records_for_report = record
        self.api_analysis_record = self.create_analysis(record)

    """
    Fnction Will write record in result excel file, it will read dict object and write data into excel file
    """
    def write_report(self):
        total_sheets = len(self.records_for_report)
        
        i = 0
        for sheet in self.records_for_report:
            self.rb.create_sheet(index=i, title=sheet)
            #set sheet active with index
            self.rb.active = i
            ws = self.rb.active
            #set Sheet tab colour
            ws.sheet_properties.tabColor = EXECUTED_SHEET_COLOR
            #append compelte tuple record into sheet row
            row_id = 1
            for row in self.records_for_report[sheet]:
                #extended the test result headers
                if row_id == 1:
                    l = list(row)
                    resultHeaders = ["ActualResponseCode","ActualResponse","TestResult"]
                    l.extend(resultHeaders)
                    row = tuple(l)
                #insert records in sheet row
                ws.append(row)

                # Give width to excel cells
                ws.row_dimensions[row_id+1].height = 200
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 40
                ws.column_dimensions['C'].width = 40
                ws.column_dimensions['D'].width = 20
                ws.column_dimensions['E'].width = 20
                ws.column_dimensions['F'].width = 40
                ws.column_dimensions['G'].width = 10
                row_id = row_id + 1
            i = i+1
        #save the test result excel sheet    
        self.rb.save(TEST_RESULTS)

        #call format file function
        format_output_file(self.api_analysis_record)

        logging.info("Validation task completed check results in  {}".format(TEST_RESULTS))
    

    """ 
    Function create data report that will be used in Graphs 
    """
    def create_analysis(self,records):
        chart_record_list = []
        for key, record in records.items():
            combineList = []
            for i in record:
                #all records merged in single tuple
                combineList.extend(i)
            
            # calulating records with tuple count function
            api_name = key #Api name
            total_test_cases = len(record)-1 #minus 1 means removing first row of heading
            total_test_case_executed = combineList.count('Y')
            total_test_case_skipped = combineList.count('N')
            total_passed = combineList.count(PASS_CELL_VALUE)
            total_failed = combineList.count(FAILED_CELL_VALUE)
            api_analysis = [api_name,total_test_cases,total_test_case_executed,total_test_case_skipped,total_passed,total_failed]    
            # adding above list records in list    
            chart_record_list.append(api_analysis)

        # return record for chart 
        # Headings [APIName,Total_testcases, testcases_executed,totaltest_skiped,passed,failed]
        # output example [['AddCustomerAPI-POST', 3, 3, 0, 3, 0], ['AddUser-POST', 2, 2, 0, 2, 0]]   
        return chart_record_list
                

    
