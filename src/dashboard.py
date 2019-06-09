from openpyxl import load_workbook
import openpyxl
from src.read_settings import *
from openpyxl.styles import Border, Side, PatternFill, Font, GradientFill, Alignment
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.drawing.fill import SolidColorFillProperties
from openpyxl.chart import (
    LineChart,
    BarChart,
    Reference,
    Series,
)
from openpyxl.drawing.fill import PatternFillProperties, ColorChoice

def format_output_file(chart_records):
    read_wb = load_workbook(filename=TEST_RESULTS)
    logging.info("Compile graphs and reports=> %s", read_wb)
    total_sheets = len(read_wb.sheetnames)

    # Css Style
    thin = Side(border_style="thin", color="000000")
    fontsize = Font(size=EXCELSHEET_FONT_SIZE)
    for sheet in range(total_sheets):
        read_wb.active = sheet
        ws = read_wb.active
        sheetRange = ws.calculate_dimension()

        # set sheet active with index
        for col in ws.iter_cols(min_row=1, max_col=ws.max_column, max_row=ws.max_row):
            for cell in col:
                # Apply excel styles for formating output
                # Apply border on excel cells
                cell.alignment = Alignment(
                    horizontal="left", vertical="center", shrinkToFit=True, wrapText=True)
                cell.border = Border(top=thin, left=thin,
                                     right=thin, bottom=thin)
                # Apply font size on all rows of excel sheet
                cell.font = fontsize
                # apply Color on cells w.r.t pass or failed cases
                if cell.value == PASS_CELL_VALUE:
                    cell.fill = PatternFill(
                        start_color=PASSED_CELL_COLOR, fill_type="solid")
                    cell.font = Font(bold=True)
                elif cell.value == FAILED_CELL_VALUE:
                    cell.fill = PatternFill(
                        start_color=FAILED_CELL_COLOR, fill_type="solid")
                    cell.font = Font(bold=True)

    # Dashboard sheet styling
    sheets = read_wb.sheetnames
    if 'Sheet' in sheets:
        read_wb.active = sheets.index(sheets[-1])
        ws = read_wb.active

        # Dashboard heading style
        ws.merge_cells('A1:Q2')
        cell = ws['A1']
        cell.value = "API Automation Test Results"
        cell.fill = PatternFill(start_color='35C4F1', fill_type="solid")
        cell.font = Font(bold=True, size=16)
        cell.alignment = Alignment(horizontal="left", vertical="center")

        # print API analysis records in table
       

        # add column headings
        #ws.append(["APINames","TotalTestCases","Executed","Skipped","Passed","Failed"])
        # output example [['BalanceInquiry-POST', 3, 3, 0, 3, 0], ['TransactionInquiry-POST', 2, 2, 0, 2, 0]] 
        
        ws.cell(row=API_TABLE_STARTS_FROM_ROW, column=1).value = "APINames"
        ws.cell(row=API_TABLE_STARTS_FROM_ROW, column=2).value = "TotalTestCases"
        ws.cell(row=API_TABLE_STARTS_FROM_ROW, column=3).value = "Executed"
        ws.cell(row=API_TABLE_STARTS_FROM_ROW, column=4).value = "Skipped"
        ws.cell(row=API_TABLE_STARTS_FROM_ROW, column=5).value = "Passed"
        ws.cell(row=API_TABLE_STARTS_FROM_ROW, column=6).value = "Failed"

        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
        ws.column_dimensions['D'].width = 15
        ws.column_dimensions['E'].width = 15
        ws.column_dimensions['F'].width = 15

        # Adding 1 beause API_TABLE_STARTS_FROM_ROW takes headings row,hence data will inserted from next row
        rows = API_TABLE_STARTS_FROM_ROW + 1
        
        for api_record in chart_records:
            columns = 1
            for val in api_record:
                ws.cell(row=rows, column=columns).value = val
                columns = columns + 1
            rows = rows + 1    

        total_record = API_TABLE_STARTS_FROM_ROW + total_sheets
        refs = 'A' + str(API_TABLE_STARTS_FROM_ROW) + ":F" + str(total_record)
        
        tab = Table(displayName="Table1", ref=refs)

        # Add a default style with striped rows and banded columns
        style = TableStyleInfo(name="TableStyleMedium9", showFirstColumn=False,
                            showLastColumn=False, showRowStripes=True, showColumnStripes=True)
        tab.tableStyleInfo = style
        ws.add_table(tab)

        #Add First chart/graph
        chart1 = BarChart()
        chart1.type = "col"
        chart1.style = 10
        chart1.height = 8 # default is 7.5
        chart1.width = 30 # default is 15
        chart1.title = "Bar Chart"
        chart1.y_axis.title = 'Total TestCases'
        chart1.x_axis.title = 'API Names'
        
        #add Category as API name
        apis = Reference(ws, min_col=1, min_row=int(API_TABLE_STARTS_FROM_ROW+1), max_row=total_record)
        data = Reference(ws, min_col=3, min_row=int(API_TABLE_STARTS_FROM_ROW), max_row=total_record,max_col=4)
        chart1.add_data(data,titles_from_data=True)
        chart1.set_categories(apis)
        chart1.shape = 4
        chart1.title = 'API TestCase Stats'
       
        ws.add_chart(chart1, "A4")
        
        #Second chart
        chart2 = BarChart()
        chart2.type = "col"
        chart2.style = 13
        chart2.height = 8 # default is 7.5
        chart2.width = 30 # default is 15
        chart2.grouping = "percentStacked"
        chart2.title = "API Success Stats"
        chart2.y_axis.title = 'Success Rate'
        chart2.overlap = 100

        #pick date from cell locations
        v2 = Reference(ws, min_col=5, min_row=int(API_TABLE_STARTS_FROM_ROW), max_col=6, max_row=total_record)
        chart2.add_data(v2, titles_from_data=True)
        chart2.set_categories(apis)

        # Style the lines
        s1 = chart2.series[0]
        s1.graphicalProperties.solidFill = "90EE90" # green
        

        s2 = chart2.series[1]
        s2.graphicalProperties.solidFill = "FF3300" # red


        #Add Chart at locton
        ws.add_chart(chart2, "A20")

    # save formated file
    read_wb.save(TEST_RESULTS)
