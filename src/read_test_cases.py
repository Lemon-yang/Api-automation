# -*- coding: utf-8 -*-
# Script details : Read Test Data from test sheet
# Created On 16-may-2019
# Develop By: Abheet Jamwal
from src.logger import logging
from src.read_settings import TESTCASES,TEST_RESULTS,API_REQUEST_COL,API_CONTEXT
from openpyxl import load_workbook,workbook

class FrameworkReader():

    def __init__(self):

        # Initialized the framework excel file 
        self.conf_testcase_file_path = TESTCASES
        self.conf_result_path = TEST_RESULTS
        self.conf_request_col = API_REQUEST_COL
        self.conf_api_context_col = API_CONTEXT
        logging.info("Reading Test Data file => %s",self.conf_testcase_file_path)
        print(self.conf_testcase_file_path)

       # Initialized the framework excel file  
        self.workbook = load_workbook(filename = self.conf_testcase_file_path)

    def _set_sheet_active(self,sheet_name,workbook):
        activeSheet = workbook.get_sheet_by_name(name = sheet_name)
        return activeSheet

   # get the all sheets name from excel workbook
    def _get_all_sheets(self,workbook):
        return workbook.get_sheet_names()

 
    def _get_total_col(self,wokbook):
        total_col = workbook.max_column
        return total_col    

    def get_testdata(self):

        #bring all sheets from testdata excel sheet
        sheets_list = self._get_all_sheets(self.workbook)
        logging.info("Reading testcase from following sheets: {}".format(str(sheets_list))) 
        # Create data dictionary where sheetName act as Key and each row is act as values
        data_dict = {}
        
        for sheet_name in sheets_list:
            ws = self._set_sheet_active(sheet_name,self.workbook)
            record_list = []
            for excel_row in ws.iter_rows(min_row=1, max_col=ws.max_column, max_row=ws.max_row, values_only=True):
                 
                # capturing API context path directly from location conf_api_context_col
                api_context = ws[self.conf_api_context_col].value
                #append api_context in list
                convertTuple = list(excel_row)
                convertTuple.insert(4,api_context)
                updatedTuple = tuple(convertTuple)
                record_list.append(updatedTuple)
            data_dict[sheet_name] = record_list
        return data_dict
                
                
                



        


        